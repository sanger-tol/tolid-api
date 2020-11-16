from openpyxl import load_workbook
import re
import requests
import sys
import os
from swagger_server.model import db, PnaUser, PnaSpecies, PnaSpecimen
from swagger_server.db_utils import create_new_specimen

def clean_cell(value):
    if value is None:
        return None
    value = str(value)
    value = re.sub("\s+", ' ', value)
    value = re.sub("\s+$", '', value)
    value = re.sub("^\s+", '', value)
    return value

def find_columns(sheet, species_column_heading):
    row = sheet[1]
    for cell in row:
        column_name = cell.value
        if column_name is None:
            break
        if (re.search(r"(?i)^taxon id$", column_name) or re.search(r"(?i)^taxon_id$", column_name) or re.search(r"(?i)^host_taxon_id$", column_name)):
            taxon_id_column = cell.column
        if (re.search(r"(?i)^specimen_id$", column_name) or re.search(r"(?i)^donor id$", column_name) or re.search(r"(?i)^donor_id$", column_name)):
            specimen_id_column = cell.column
        if (re.search(r"(?i)^"+species_column_heading+"$", column_name)):
            scientific_name_column = cell.column
        if (re.search(r"(?i)^public_name$", column_name)):
            public_name_column = cell.column

    return (taxon_id_column, specimen_id_column, scientific_name_column, public_name_column)

def validate_excel(dirname=None, filename=None, user=None, species_column_heading=None):
    workbook = load_workbook(filename=dirname+'/'+filename)
    sheet = workbook.active
    errors = []
    validated = validate_sheet(sheet, assign=False, user=user, species_column_heading=species_column_heading, errors=errors)
    if validated:
        # Can go through and assign public names
        validate_sheet(sheet, assign=True, user=user, species_column_heading=species_column_heading)
        new_filename = re.sub(r"\.xlsx$", '-validated.xlsx', filename)
        workbook.save(dirname+'/'+new_filename)
        return True, new_filename, errors
    else:
        return False, None, errors

def validate_sheet(sheet, assign=False, user=None, species_column_heading=None, errors=[]):
    ok = True
    (taxon_id_column, specimen_id_column, scientific_name_column, public_name_column) = find_columns(sheet, species_column_heading)
    current_row = 2 
    for row in sheet.iter_rows(min_row=current_row, max_row=sheet.max_row, values_only=True):
        taxon_id = clean_cell(row[taxon_id_column-1])
        specimen_id = clean_cell(row[specimen_id_column-1])
        scientific_name = clean_cell(row[scientific_name_column-1])
        if (taxon_id is None) or (specimen_id is None) or (scientific_name is None):
            break
        if (assign):
            existing_public_name = sheet.cell(row=current_row, column=public_name_column).value
            if (existing_public_name is None):
                existing_specimen = db.session.query(PnaSpecimen).filter(PnaSpecimen.specimen_id == specimen_id).filter(PnaSpecimen.species_id == taxon_id).one_or_none()
                if (existing_specimen is not None):
                    sheet.cell(row=current_row, column=public_name_column, value=existing_specimen.public_name)
                else:
                    existing_species = db.session.query(PnaSpecies).filter(PnaSpecies.taxonomy_id == taxon_id).one_or_none()
                    new_specimen = create_new_specimen(existing_species, specimen_id, user)
                    db.session.add(new_specimen)
                    db.session.commit()
                    sheet.cell(row=current_row, column=public_name_column, value=new_specimen.public_name)
        else:
            if (re.search(r"sp\.$", scientific_name)):
                errors.append({"message": "Row "+str(current_row)+": Genus only for "+scientific_name+", not assigning public name"})
                ok = False
            else:
                # Search for the taxomomy ID
                existing_species = db.session.query(PnaSpecies).filter(PnaSpecies.taxonomy_id == taxon_id).one_or_none()
                if existing_species is None:
                    errors.append({"message": "Row "+str(current_row)+": Taxon ID " + taxon_id + " cannot be found"})
                    ok = False
                else:
                    if (scientific_name != existing_species.name):
                        errors.append({"message": "Row "+str(current_row)+": Expecting "+scientific_name+", got "+ existing_species.name})
                        ok = False
                    else:
                        # Search for the public name
                        existing_specimen = db.session.query(PnaSpecimen).filter(PnaSpecimen.specimen_id == specimen_id).filter(PnaSpecimen.species_id == taxon_id).one_or_none()
                        if (existing_specimen is not None):
                            sheet.cell(row=current_row, column=public_name_column, value=existing_specimen.public_name)
                        else:
                            # No existing public name - deal with this later - nothing to do now
                            pass
        current_row+=1
    return ok

