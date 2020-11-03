# Script to validate a manifest and fill in the public name column
# using the public_name_api service
#
# Needs PUBLIC_NAME_API_URL and PUBLIC_NAME_API_KEY setting before running
# env $(cat .env.uat | xargs) python validate_excel.py test.xlsx dtol

from openpyxl import load_workbook
import re
import requests
import sys
import os

public_name_api_url = os.environ["PUBLIC_NAME_API_URL"]
public_name_api_key = os.environ["PUBLIC_NAME_API_KEY"]

filename = sys.argv[1]
project = sys.argv[2]

def clean_cell(value):
    if value is None:
        return None
    value = str(value)
    value = re.sub("\s+", ' ', value)
    value = re.sub("\s+$", '', value)
    value = re.sub("^\s+", '', value)
    return value

def find_columns(sheet):
    # What is the taxonomy ID column?
    row = sheet[1]
    for cell in row:
        column_name = cell.value
        if column_name is None:
            break
        if (re.search(r"(?i)^taxon id$", column_name) or re.search(r"(?i)^taxon_id$", column_name) or re.search(r"(?i)^host_taxon_id$", column_name)):
            taxon_id_column = cell.column
        if (re.search(r"(?i)^specimen_id$", column_name) or re.search(r"(?i)^donor id$", column_name) or re.search(r"(?i)^donor_id$", column_name)):
            specimen_id_column = cell.column
        if (re.search(r"(?i)^scientific_name$", column_name) or (re.search(r"(?i)^common_name$", column_name) and (project=="sanger")) or (re.search(r"(?i)^host_scientific_name$", column_name) and (project=="asg"))):
            scientific_name_column = cell.column
        if (re.search(r"(?i)^public_name$", column_name)):
            public_name_column = cell.column

    return (taxon_id_column, specimen_id_column, scientific_name_column, public_name_column)

def validate_excel(filename):
    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    validated = validate_sheet(sheet, assign=False)
    if validated:
        # Can go through and assign public names
        validate_sheet(sheet, assign=True)
        new_filename = re.sub(r"\.xlsx$", '-validated.xlsx', filename)
        print("ALL CORRECT. SAVED AS " + new_filename)
        workbook.save(new_filename)
    else:
        print("VALIDATION FAILURES. SEE MESSAGES ABOVE")

def validate_sheet(sheet, assign=False):
    ok = True
    (taxon_id_column, specimen_id_column, scientific_name_column, public_name_column) = find_columns(sheet)
    current_row = 2 
    for row in sheet.iter_rows(min_row=current_row, max_row=sheet.max_row, values_only=True):
        taxon_id = clean_cell(row[taxon_id_column-1])
        specimen_id = clean_cell(row[specimen_id_column-1])
        scientific_name = clean_cell(row[scientific_name_column-1])
        if (taxon_id is None) or (specimen_id is None) or (scientific_name is None):
            break
        if (assign):
            existing_public_name = sheet.cell(row=current_row, column=public_name_column).value
            if (existing_public_name is None):
                print("ASSIGNING NEW PUBLIC NAME FOR " + str(taxon_id)+"-"+str(specimen_id)+" ("+str(scientific_name)+")" )
                r = requests.put(public_name_api_url+'/public-name', params={'specimenId': specimen_id, 'taxonomyId': taxon_id}, headers={'api-key': public_name_api_key})
                if (r.status_code == 200):
                    json = r.json()
                    new_public_name = json[0]["publicName"]
                    sheet.cell(row=current_row, column=public_name_column, value=new_public_name)
                else:
                    print("ERROR: "+ r.text)
        else:
            print("Checking: "+str(taxon_id)+"-"+str(specimen_id)+" ("+str(scientific_name)+")")
            if (re.search(r"sp\.$", scientific_name)):
                print("ERROR: Genus only for "+scientific_name+", not assigning public name")
                ok = False
            else:
                r = requests.get(public_name_api_url+'/public-name', params={'specimenId': specimen_id, 'taxonomyId': taxon_id})
                json = r.json()
                if (r.status_code == 200):
                    if (len(json) > 0):
                        if (scientific_name != json[0]["species"]):
                            print("ERROR: Expecting "+ scientific_name + " GOT "+ json[0]["species"])
                            ok = False
                            next

                        existing_public_name = json[0]["publicName"]
                        sheet.cell(row=current_row, column=public_name_column, value=existing_public_name)
                    else:
                        # No existing public name - deal with this later
                        pass
                else:
                    # Failure status code
                    print("ERROR: "+ r.text)
                    ok = False
        current_row+=1
    return ok

validate_excel(filename)
