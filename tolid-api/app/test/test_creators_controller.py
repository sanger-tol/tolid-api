# SPDX-FileCopyrightText: 2021 Genome Research Ltd.
#
# SPDX-License-Identifier: MIT

from __future__ import absolute_import

from test import BaseTestCase

from main.excel_utils import find_columns

from openpyxl import load_workbook


class TestCreatorsController(BaseTestCase):

    def test_add_tolid(self):
        # No authorisation token given
        query_string = []
        response = self.client.open(
            '/api/v2/tol-ids',
            method='PUT',
            query_string=query_string)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        # Invalid authorisation token given
        query_string = []
        response = self.client.open(
            '/api/v2/tol-ids',
            method='PUT',
            headers={'api-key': '12345678'},
            query_string=query_string)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # No taxonomyId given
        query_string = []
        response = self.client.open(
            '/api/v2/tol-ids',
            method='PUT',
            headers={'api-key': self.user3.api_key},
            query_string=query_string)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # No specimenId given
        query_string = [('taxonomyId', 6344)]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='PUT',
            headers={'api-key': self.user3.api_key},
            query_string=query_string)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Taxonomy ID not in database
        query_string = [('taxonomyId', 999999999),
                        ('specimenId', 'SAN0000100')]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='PUT',
            headers={'api-key': self.user3.api_key},
            query_string=query_string)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # User not a creator
        query_string = [('taxonomyId', '6355'),
                        ('specimenId', 'SAN0000100')]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='PUT',
            headers={'api-key': self.user1.api_key},
            query_string=query_string)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Second taxonomy ID for specimen
        query_string = [('taxonomyId', '6355'),
                        ('specimenId', 'SAN0000100')]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='PUT',
            headers={'api-key': self.user3.api_key},
            query_string=query_string)
        expect = [{
            'species': {
                'commonName': 'None',
                'currentHighestTolidNumber': 2,
                'family': 'Nereididae',
                'genus': 'Perinereis',
                'kingdom': 'Metazoa',
                'order': 'Phyllodocida',
                'phylum': 'Annelida',
                'prefix': 'wpPerVanc',
                'scientificName': 'Perinereis vancaurica',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6355
            },
            'tolId': 'wpPerVanc2',
            'specimen': {'specimenId': 'SAN0000100'},
        }]
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEqual(expect, response.json)

        # Specimen ID not in database - should create it
        query_string = [('taxonomyId', 6344),
                        ('specimenId', 'SAN0000100xxxxx')]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='PUT',
            headers={'api-key': self.user3.api_key},
            query_string=query_string)
        expect = [{
            'species': {
                'commonName': 'lugworm',
                'currentHighestTolidNumber': 3,
                'family': 'Arenicolidae',
                'genus': 'Arenicola',
                'order': 'None',
                'phylum': 'Annelida',
                'kingdom': 'Metazoa',
                'prefix': 'wuAreMari',
                'scientificName': 'Arenicola marina',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6344
            },
            'tolId': 'wuAreMari3',
            'specimen': {'specimenId': 'SAN0000100xxxxx'},
        }]
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEqual(expect, response.json)

        # Specimen ID not in database and first for species - should create it
        query_string = [('taxonomyId', 9606),
                        ('specimenId', 'SAN0000999xxxxx')]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='PUT',
            headers={'api-key': self.user3.api_key},
            query_string=query_string)
        expect = [{
            'species': {
                'commonName': 'human',
                'currentHighestTolidNumber': 1,
                'family': 'Hominidae',
                'genus': 'Homo',
                'order': 'Primates',
                'phylum': 'Chordata',
                'kingdom': 'Metazoa',
                'prefix': 'mHomSap',
                'scientificName': 'Homo sapiens',
                'taxaClass': 'Mammalia',
                'taxonomyId': 9606
            },
            'tolId': 'mHomSap1',
            'specimen': {'specimenId': 'SAN0000999xxxxx'}
        }]
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEqual(expect, response.json)

        # Existing - should return existing
        query_string = [('taxonomyId', 6344),
                        ('specimenId', 'SAN0000100')]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='PUT',
            headers={'api-key': self.user3.api_key},
            query_string=query_string)
        expect = [{
            'species': {
                'commonName': 'lugworm',
                'currentHighestTolidNumber': 3,
                'family': 'Arenicolidae',
                'genus': 'Arenicola',
                'order': 'None',
                'phylum': 'Annelida',
                'kingdom': 'Metazoa',
                'prefix': 'wuAreMari',
                'scientificName': 'Arenicola marina',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6344
            },
            'tolId': 'wuAreMari1',
            'specimen': {'specimenId': 'SAN0000100'},
        }]
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEqual(expect, response.json)

    def test_bulk_search_tol_ids(self):
        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v2/tol-ids',
            method='POST',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v2/tol-ids',
            method='POST',
            headers={'api-key': '12345678'},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # No taxonomyId given
        body = [{}]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='POST',
            headers={'api-key': self.user3.api_key},
            json=body)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # No specimenId given
        body = [{'taxonomyId': 6344}]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='POST',
            headers={'api-key': self.user3.api_key},
            json=body)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Taxonomy ID not in database - should get a request back
        body = [{'taxonomyId': 999999999,
                 'specimenId': 'SAN0000100'}]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='POST',
            headers={'api-key': self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expect = [{'createdBy': {'email': 'test_user_creator@sanger.ac.uk',
                                 'name': 'test_user_creator',
                                 'organisation': 'Sanger Institute',
                                 'roles': [{'role': 'creator'}]},
                   'reason': None,
                   'requestId': 1,
                   'species': {'taxonomyId': 999999999},
                   'specimen': {'specimenId': 'SAN0000100'},
                   'status': 'Pending'}]
        self.assertEqual(expect, response.json)

        # User doesn't have creator role
        body = [{'taxonomyId': 6344,
                'specimenId': 'SAN0000100'}]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='POST',
            headers={'api-key': self.user1.api_key},
            json=body)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Specimen ID not in database, multiple taxons for same specimen - should create them
        body = [{'taxonomyId': 6344,
                'specimenId': 'SAN0000100xxxxx'},
                {'taxonomyId': 6355,
                'specimenId': 'SAN0000100xxxxx'}]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='POST',
            headers={'api-key': self.user3.api_key},
            json=body)
        expect = [{
            'species': {
                'commonName': 'lugworm',
                'currentHighestTolidNumber': 3,
                'family': 'Arenicolidae',
                'genus': 'Arenicola',
                'order': 'None',
                'phylum': 'Annelida',
                'kingdom': 'Metazoa',
                'prefix': 'wuAreMari',
                'scientificName': 'Arenicola marina',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6344
            },
            'tolId': 'wuAreMari3',
            'specimen': {'specimenId': 'SAN0000100xxxxx'},
        }, {
            'species': {
                'commonName': 'None',
                'currentHighestTolidNumber': 2,
                'family': 'Nereididae',
                'genus': 'Perinereis',
                'kingdom': 'Metazoa',
                'order': 'Phyllodocida',
                'phylum': 'Annelida',
                'prefix': 'wpPerVanc',
                'scientificName': 'Perinereis vancaurica',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6355
            },
            'tolId': 'wpPerVanc2',
            'specimen': {'specimenId': 'SAN0000100xxxxx'},
        }]
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEqual(expect, response.json)

        # Single search for existing
        body = [{'taxonomyId': 6344,
                'specimenId': 'SAN0000100'}]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='POST',
            headers={'api-key': self.user3.api_key},
            json=body)
        expect = [{
            'species': {
                'commonName': 'lugworm',
                'currentHighestTolidNumber': 3,
                'family': 'Arenicolidae',
                'genus': 'Arenicola',
                'order': 'None',
                'phylum': 'Annelida',
                'kingdom': 'Metazoa',
                'prefix': 'wuAreMari',
                'scientificName': 'Arenicola marina',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6344
            },
            'tolId': 'wuAreMari1',
            'specimen': {'specimenId': 'SAN0000100'},
        }]
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEqual(expect, response.json)

        # Search for existing and new
        body = [{'taxonomyId': 6344,
                'specimenId': 'SAN0000100'},
                {'taxonomyId': 6344,
                'specimenId': 'SAN0000100wwwww'}]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='POST',
            headers={'api-key': self.user3.api_key},
            json=body)
        expect = [{
            'species': {
                'commonName': 'lugworm',
                'currentHighestTolidNumber': 4,
                'family': 'Arenicolidae',
                'genus': 'Arenicola',
                'order': 'None',
                'phylum': 'Annelida',
                'kingdom': 'Metazoa',
                'prefix': 'wuAreMari',
                'scientificName': 'Arenicola marina',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6344
            },
            'tolId': 'wuAreMari1',
            'specimen': {'specimenId': 'SAN0000100'},
        }, {
            'species': {
                'commonName': 'lugworm',
                'currentHighestTolidNumber': 4,
                'family': 'Arenicolidae',
                'genus': 'Arenicola',
                'order': 'None',
                'phylum': 'Annelida',
                'kingdom': 'Metazoa',
                'prefix': 'wuAreMari',
                'scientificName': 'Arenicola marina',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6344
            },
            'tolId': 'wuAreMari4',
            'specimen': {'specimenId': 'SAN0000100wwwww'},
        }]
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEqual(expect, response.json)

        # Search for existing and 2 new
        body = [{'taxonomyId': 6344,
                'specimenId': 'SAN0000100'},
                {'taxonomyId': 6344,
                'specimenId': 'SAN0000100ppppp'},
                {'taxonomyId': 6344,
                'specimenId': 'SAN0000100qqqqq'}]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='POST',
            headers={'api-key': self.user3.api_key},
            json=body)
        expect = [{
            'species': {
                'commonName': 'lugworm',
                'currentHighestTolidNumber': 6,
                'family': 'Arenicolidae',
                'genus': 'Arenicola',
                'order': 'None',
                'phylum': 'Annelida',
                'kingdom': 'Metazoa',
                'prefix': 'wuAreMari',
                'scientificName': 'Arenicola marina',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6344
            },
            'tolId': 'wuAreMari1',
            'specimen': {'specimenId': 'SAN0000100'},
        }, {
            'species': {
                'commonName': 'lugworm',
                'currentHighestTolidNumber': 6,
                'family': 'Arenicolidae',
                'genus': 'Arenicola',
                'order': 'None',
                'phylum': 'Annelida',
                'kingdom': 'Metazoa',
                'prefix': 'wuAreMari',
                'scientificName': 'Arenicola marina',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6344
            },
            'tolId': 'wuAreMari5',
            'specimen': {'specimenId': 'SAN0000100ppppp'},
        }, {
            'species': {
                'commonName': 'lugworm',
                'currentHighestTolidNumber': 6,
                'family': 'Arenicolidae',
                'genus': 'Arenicola',
                'order': 'None',
                'phylum': 'Annelida',
                'kingdom': 'Metazoa',
                'prefix': 'wuAreMari',
                'scientificName': 'Arenicola marina',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6344
            },
            'tolId': 'wuAreMari6',
            'specimen': {'specimenId': 'SAN0000100qqqqq'},
        }]
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEqual(expect, response.json)

        # Search for existing and new, plus duplicated queries
        body = [{'taxonomyId': 6344,
                'specimenId': 'SAN0000100'},
                {'taxonomyId': 6344,
                'specimenId': 'SAN0000100rrrrr'},
                {'taxonomyId': 6344,
                'specimenId': 'SAN0000100'},
                {'taxonomyId': 6344,
                'specimenId': 'SAN0000100rrrrr'}]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='POST',
            headers={'api-key': self.user3.api_key},
            json=body)
        expect = [{
            'species': {
                'commonName': 'lugworm',
                'currentHighestTolidNumber': 7,
                'family': 'Arenicolidae',
                'genus': 'Arenicola',
                'order': 'None',
                'phylum': 'Annelida',
                'kingdom': 'Metazoa',
                'prefix': 'wuAreMari',
                'scientificName': 'Arenicola marina',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6344
            },
            'tolId': 'wuAreMari1',
            'specimen': {'specimenId': 'SAN0000100'},
        }, {
            'species': {
                'commonName': 'lugworm',
                'currentHighestTolidNumber': 7,
                'family': 'Arenicolidae',
                'genus': 'Arenicola',
                'order': 'None',
                'phylum': 'Annelida',
                'kingdom': 'Metazoa',
                'prefix': 'wuAreMari',
                'scientificName': 'Arenicola marina',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6344
            },
            'tolId': 'wuAreMari7',   # We created wuAreMari3,4,5,6 earlier on in this method
            'specimen': {'specimenId': 'SAN0000100rrrrr'},
        }, {
            'species': {
                'commonName': 'lugworm',
                'currentHighestTolidNumber': 7,
                'family': 'Arenicolidae',
                'genus': 'Arenicola',
                'order': 'None',
                'phylum': 'Annelida',
                'kingdom': 'Metazoa',
                'prefix': 'wuAreMari',
                'scientificName': 'Arenicola marina',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6344
            },
            'tolId': 'wuAreMari1',
            'specimen': {'specimenId': 'SAN0000100'},
        }, {
            'species': {
                'commonName': 'lugworm',
                'currentHighestTolidNumber': 7,
                'family': 'Arenicolidae',
                'genus': 'Arenicola',
                'order': 'None',
                'phylum': 'Annelida',
                'kingdom': 'Metazoa',
                'prefix': 'wuAreMari',
                'scientificName': 'Arenicola marina',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6344
            },
            'tolId': 'wuAreMari7',
            'specimen': {'specimenId': 'SAN0000100rrrrr'},
        }]

        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEqual(expect, response.json)

        # Later query returns a request
        body = [{'taxonomyId': 6344,
                'specimenId': 'SAN0000100bbbbb'},
                {'taxonomyId': 9999999,
                'specimenId': 'SAN0000100'}]
        response = self.client.open(
            '/api/v2/tol-ids',
            method='POST',
            headers={'api-key': self.user3.api_key},
            json=body)

        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expect = [{
            'species': {
                'commonName': 'lugworm',
                'currentHighestTolidNumber': 8,
                'family': 'Arenicolidae',
                'genus': 'Arenicola',
                'order': 'None',
                'phylum': 'Annelida',
                'kingdom': 'Metazoa',
                'prefix': 'wuAreMari',
                'scientificName': 'Arenicola marina',
                'taxaClass': 'Polychaeta',
                'taxonomyId': 6344
            },
            'tolId': 'wuAreMari8',
            'specimen': {'specimenId': 'SAN0000100bbbbb'},
        }, {
            'createdBy': {'email': 'test_user_creator@sanger.ac.uk',
                          'name': 'test_user_creator',
                          'organisation': 'Sanger Institute',
                          'roles': [{'role': 'creator'}]},
            'reason': None,
            'requestId': 2,
            'species': {'taxonomyId': 9999999},
            'specimen': {'specimenId': 'SAN0000100'},
            'status': 'Pending'}]
        self.assertEqual(expect, response.json)

        # And the new one should have been inserted
        response = self.client.open(
            '/api/v2/specimens/SAN0000100bbbbb',
            method='GET')
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expect = [{
            'specimenId': 'SAN0000100bbbbb',
            'tolIds': [
                {'species': {'commonName': 'lugworm',
                             'currentHighestTolidNumber': 8,
                             'family': 'Arenicolidae',
                             'genus': 'Arenicola',
                             'order': 'None',
                             'phylum': 'Annelida',
                             'kingdom': 'Metazoa',
                             'prefix': 'wuAreMari',
                             'scientificName': 'Arenicola marina',
                             'taxaClass': 'Polychaeta',
                             'taxonomyId': 6344},
                 'tolId': 'wuAreMari8'}]
        }]
        self.assertEqual(expect, response.json)

        response = self.client.open(
            '/api/v2/requests/2',
            method='GET')
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expect = [{'createdBy': {'email': 'test_user_creator@sanger.ac.uk',
                                 'name': 'test_user_creator',
                                 'organisation': 'Sanger Institute',
                                 'roles': [{'role': 'creator'}]},
                   'reason': None,
                   'requestId': 2,
                   'species': {'taxonomyId': 9999999},
                   'specimen': {'specimenId': 'SAN0000100'},
                   'status': 'Pending'}]
        self.assertEqual(expect, response.json)

    def test_validate_manifest(self):
        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v2/validate-manifest',
            method='POST',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v2/validate-manifest',
            method='POST',
            headers={'api-key': '12345678'},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Excel file missing
        data = {}
        response = self.client.open(
            '/api/v2/validate-manifest',
            method='POST',
            headers={'api-key': self.user1.api_key},
            data=data)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Excel file with no taxon ID, specimen ID, ToLID column
        file = open('test/test-manifest-no-columns.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'),
        }
        expected = {'errors': [{'message': 'Cannot find Taxon ID column'},
                               {'message': 'Cannot find Specimen ID column'},
                               {'message': 'Cannot find ToLID column'}]}
        response = self.client.open(
            '/api/v2/validate-manifest',
            method='POST',
            headers={'api-key': self.user3.api_key},
            data=data)
        file.close()
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEqual(expected, response.json)

        # Excel file with errors
        file = open('test/test-manifest-with-errors.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'),
        }
        expected = {'errors':
                    [{'message': 'Row 2: Expecting Arenicola marina, got Homo sapiens'},
                     {'message': 'Row 3: Taxon ID 9999999 cannot be found'},
                     {'message': 'Row 4: Genus only for Arenicola sp., not assigning ToLID'}]}
        response = self.client.open(
            '/api/v2/validate-manifest',
            method='POST',
            headers={'api-key': self.user3.api_key},
            data=data)
        file.close()
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEqual(expected, response.json)

        # User not a creator
        file = open('test/test-manifest.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'),
        }
        response = self.client.open(
            '/api/v2/validate-manifest',
            method='POST',
            headers={'api-key': self.user1.api_key},
            data=data)
        file.close()
        self.assert403(response, 'Not received a 403 response')

        # Excel file correct
        file = open('test/test-manifest.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'),
        }
        response = self.client.open(
            '/api/v2/validate-manifest',
            method='POST',
            headers={'api-key': self.user3.api_key},
            data=data)
        file.close()
        self.assert200(response, 'Not received a 200 response')
        self.assertEqual('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         response.content_type)

        # Save as Excel file
        file = open('test/test-manifest-validated.xlsx', 'wb')
        file.write(response.get_data())
        file.close()
        workbook = load_workbook(filename='test/test-manifest-validated.xlsx')
        sheet = workbook.active
        (taxon_id_column, specimen_id_column, scientific_name_column, tol_id_column) = \
            find_columns(sheet, 'scientific_name')
        self.assertEqual('wuAreMari3', sheet.cell(row=2, column=tol_id_column).value)
        self.assertEqual('wuAreMari4', sheet.cell(row=3, column=tol_id_column).value)
        self.assertEqual('wuAreMari4', sheet.cell(row=4, column=tol_id_column).value)

        # Different column name for species
        file = open('test/test-manifest-col.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'),
        }
        query_string = {'speciesColumnHeading': 'random column name'}
        response = self.client.open(
            '/api/v2/validate-manifest',
            method='POST',
            headers={'api-key': self.user3.api_key},
            query_string=query_string,
            data=data)
        file.close()
        self.assert200(response, 'Not received a 200 response')
        self.assertEqual('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         response.content_type)

        # Save as Excel file
        file = open('test/test-manifest-validated.xlsx', 'wb')
        file.write(response.get_data())
        file.close()
        workbook = load_workbook(filename='test/test-manifest-validated.xlsx')
        sheet = workbook.active
        (taxon_id_column, specimen_id_column, scientific_name_column, tol_id_column) = \
            find_columns(sheet, 'random column name')
        self.assertEqual('wuAreMari3', sheet.cell(row=2, column=tol_id_column).value)
        self.assertEqual('wuAreMari4', sheet.cell(row=3, column=tol_id_column).value)


if __name__ == '__main__':
    import unittest
    unittest.main()
